#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pure game logic — no tkinter dependency.
Shared between server.py and any tests.
"""

import json
import io
import os
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─── Constants ────────────────────────────────────────────────────────────────
GRAPH: Dict[Any, List[Any]] = {
    "hub": [1, 2, 3],
    1: ["hub", 2, 4, 5],
    2: ["hub", 1, 3, 6],
    3: ["hub", 2, 7, 8],
    4: [1, 5, 9],
    5: [1, 4, 6, 10],
    6: [2, 5, 7, 11],
    7: [3, 6, 8, 12],
    8: [3, 7, 13, 14],
    9: [4, 10, 15],
    10: [5, 9, 11, 16],
    11: [6, 10, 12, 17],
    12: [7, 11, 13, 18],
    13: [8, 12, 14, 19],
    14: [8, 13, 15, 20],
    15: [9, 14, 16, 20],
    16: [10, 15, 17],
    17: [11, 16, 18],
    18: [12, 17, 19],
    19: [13, 18, 20],
    20: [14, 15, 19],
}

IMAGE_VALUES: Dict[int, int] = {
    1: 1, 2: 1, 3: 1, 4: 1,
    5: 2, 6: 2, 7: 2, 8: 2,
    9: 3, 10: 3, 11: 3, 12: 3,
    13: 4, 14: 4, 15: 4, 16: 4,
    17: 5, 18: 5, 19: 5, 20: 5,
}

IMAGE_FILES: Dict[int, str] = {i: f"{i}.jpg" for i in range(1, 21)}
IMAGE_FILES[14] = "14.png"

# Canvas positions for JS board drawing (800×500 canvas)
NODE_POSITIONS: Dict[Any, tuple] = {
    "hub": (400, 22),
    1: (142, 95),  2: (400, 95),  3: (658, 95),
    4: (49, 183),  5: (187, 183), 6: (347, 183), 7: (507, 183), 8: (658, 183),
    9: (36, 276),  10: (142, 276), 11: (253, 276), 12: (369, 276),
    13: (484, 276), 14: (604, 276),
    15: (89, 369),  16: (200, 369), 17: (316, 369), 18: (436, 369),
    19: (556, 369), 20: (671, 369),
}

PLAYER_COLORS = ["#E74C3C", "#3498DB", "#2ECC71", "#F39C12", "#9B59B6", "#1ABC9C"]

METAPHORS_TO_WIN = 5
TIMER_SECONDS = 45
BEST_METAPHOR_BONUS = 5


# ─── Data Classes ─────────────────────────────────────────────────────────────
@dataclass
class MetaphorRecord:
    player_name: str
    metaphor_text: str
    images_used: List[int]
    score: int
    round_number: int
    is_best: bool = False
    votes: int = 0
    metaphor_id: int = 0

    def to_dict(self):
        return {
            "id": self.metaphor_id,
            "player_name": self.player_name,
            "metaphor_text": self.metaphor_text,
            "images_used": self.images_used,
            "score": self.score,
            "round_number": self.round_number,
            "is_best": self.is_best,
            "votes": self.votes,
        }


@dataclass
class Player:
    name: str
    sid: str = ""
    color: str = "#E74C3C"
    current_node: Any = "hub"
    metaphor_count: int = 0
    total_score: int = 0
    metaphors: List[MetaphorRecord] = field(default_factory=list)

    def to_dict(self):
        return {
            "name": self.name,
            "color": self.color,
            "current_node": self.current_node,
            "metaphor_count": self.metaphor_count,
            "total_score": self.total_score,
        }


class GameState:
    def __init__(self, player_names: List[str], player_sids: List[str]):
        self.players = [
            Player(name=n, sid=s, color=PLAYER_COLORS[i % len(PLAYER_COLORS)])
            for i, (n, s) in enumerate(zip(player_names, player_sids))
        ]
        self.current_player_index = 0
        self.round_number = 1
        self.all_metaphors: List[MetaphorRecord] = []
        self._next_metaphor_id = 1
        self.game_over = False
        self.winner: Optional[Player] = None
        self.phase = "move"
        self.selected_images: List[int] = []
        # Appeal state
        self.pending_metaphor: Optional[dict] = None
        self.appeal_votes: Dict[str, bool] = {}  # sid → approve/reject
        self.appeal_voter_sids: List[str] = []
        # Endgame state
        self.endgame_votes: Dict[str, int] = {}  # sid → metaphor_id
        self.timer_end: float = 0.0

    @property
    def current_player(self) -> Player:
        return self.players[self.current_player_index]

    def next_turn(self):
        self.selected_images = []
        self.current_player_index = (self.current_player_index + 1) % len(self.players)
        if self.current_player_index == 0:
            self.round_number += 1
        self.phase = "move"

    def check_winner(self) -> Optional[Player]:
        for p in self.players:
            if p.metaphor_count >= METAPHORS_TO_WIN:
                return p
        return None

    def add_metaphor(self, player: Player, text: str, images: List[int]) -> MetaphorRecord:
        score = sum(IMAGE_VALUES[i] for i in images)
        if len(images) >= 3:
            score *= 2
        rec = MetaphorRecord(
            player_name=player.name,
            metaphor_text=text,
            images_used=images,
            score=score,
            round_number=self.round_number,
            metaphor_id=self._next_metaphor_id,
        )
        self._next_metaphor_id += 1
        player.metaphors.append(rec)
        player.metaphor_count += 1
        player.total_score += score
        self.all_metaphors.append(rec)
        return rec

    def to_dict(self, include_timer=True) -> dict:
        d = {
            "phase": self.phase,
            "current_player": self.current_player.to_dict(),
            "current_player_sid": self.current_player.sid,
            "players": [p.to_dict() for p in self.players],
            "selected_images": self.selected_images,
            "metaphors": [m.to_dict() for m in self.all_metaphors],
            "round_number": self.round_number,
            "winner": self.winner.name if self.winner else None,
            "game_over": self.game_over,
        }
        if include_timer:
            d["timer_end"] = self.timer_end
        if self.pending_metaphor:
            d["pending_metaphor"] = self.pending_metaphor
            d["appeal_voted_sids"] = list(self.appeal_votes.keys())
            d["appeal_votes_approve"] = sum(1 for v in self.appeal_votes.values() if v)
            d["appeal_votes_reject"] = sum(1 for v in self.appeal_votes.values() if not v)
        if self.phase == "endgame":
            d["endgame_voted_sids"] = list(self.endgame_votes.keys())
        return d


# ─── Graph Helpers ────────────────────────────────────────────────────────────
def are_connected(nodes: List[int]) -> bool:
    """Return True if all nodes form a connected subgraph."""
    if len(nodes) <= 1:
        return True
    s = set(nodes)
    visited = {nodes[0]}
    queue = [nodes[0]]
    while queue:
        cur = queue.pop(0)
        for nb in GRAPH.get(cur, []):
            if isinstance(nb, int) and nb in s and nb not in visited:
                visited.add(nb)
                queue.append(nb)
    return visited == s


def get_adjacent(node: Any) -> List[Any]:
    return GRAPH.get(node, [])


def compute_score(images: List[int]) -> int:
    score = sum(IMAGE_VALUES[i] for i in images)
    if len(images) >= 3:
        score *= 2
    return score


# ─── Validation ───────────────────────────────────────────────────────────────
def validate_local(text: str) -> dict:
    text = text.strip()
    if len(text) < 4:
        return {"valid": False, "reason": "המטפורה קצרה מדי"}
    words = text.split()
    if len(words) < 2:
        return {"valid": False, "reason": "נדרשות לפחות 2 מילים"}
    if all(w.isdigit() for w in words):
        return {"valid": False, "reason": "המטפורה אינה יכולה להיות רק מספרים"}
    letter_count = sum(1 for c in text if c.isalpha())
    if letter_count < len(text) * 0.35:
        return {"valid": False, "reason": "המטפורה אינה מכילה מספיק אותיות"}
    return {"valid": True, "reason": "המטפורה תקינה"}


def validate_with_claude(text: str, image_ids: List[int], api_key: str) -> dict:
    try:
        client = anthropic.Anthropic(api_key=api_key)
        image_desc = ", ".join(f"תמונה #{i} (ערך {IMAGE_VALUES[i]})" for i in image_ids)
        prompt = (
            f'You are a judge in a creative metaphor board game.\n'
            f'Player wrote this metaphor: "{text}"\n'
            f'Images used: {image_desc}\n\n'
            f'A VALID metaphor: creative comparison, meaningful, at least 2 words, shows imagination.\n'
            f'INVALID only if: pure gibberish keystrokes, random letters only, completely empty.\n\n'
            f'Respond ONLY with JSON (no other text):\n'
            f'{{"valid": true/false, "reason": "brief explanation in same language as metaphor"}}'
        )
        msg = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}]
        )
        return json.loads(msg.content[0].text.strip())
    except Exception:
        return validate_local(text)


def validate_metaphor(text: str, image_ids: List[int], api_key: str) -> dict:
    if api_key and ANTHROPIC_AVAILABLE:
        return validate_with_claude(text, image_ids, api_key)
    return validate_local(text)


# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_to_excel(game_state: GameState) -> io.BytesIO:
    """Build Excel workbook and return as BytesIO buffer (for HTTP download)."""
    wb = openpyxl.Workbook()

    # Sheet 1: Metaphors
    ws1 = wb.active
    ws1.title = "מטפורות"
    try:
        ws1.sheet_view.rightToLeft = True
    except Exception:
        pass
    h1 = ["מטפורה", "ניקוד", "שם משתתף", "מספר תמונות", "סיבוב", "מטפורה מנצחת"]
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86C1")
        cell.alignment = Alignment(horizontal="center")
    for r in game_state.all_metaphors:
        ws1.append([
            r.metaphor_text,
            r.score,
            r.player_name,
            len(r.images_used),
            r.round_number,
            "✓" if r.is_best else ""
        ])
    ws1.column_dimensions["A"].width = 55
    ws1.column_dimensions["C"].width = 22
    for col in ["B", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 18

    # Sheet 2: Final Scores
    ws2 = wb.create_sheet("תוצאות סופיות")
    try:
        ws2.sheet_view.rightToLeft = True
    except Exception:
        pass
    h2 = ["שם משתתף", "ניקוד משחק", "מטפורות שנוצרו", "בונוס", "סה\"כ"]
    ws2.append(h2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A5276")
        cell.alignment = Alignment(horizontal="center")
    for p in game_state.players:
        has_best = any(r.is_best and r.player_name == p.name for r in game_state.all_metaphors)
        bonus = BEST_METAPHOR_BONUS if has_best else 0
        ws2.append([p.name, p.total_score, p.metaphor_count, bonus, p.total_score + bonus])
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
